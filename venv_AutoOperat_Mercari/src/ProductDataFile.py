import os
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles.borders import Border, Side
import sys
import ProductInfo
import Const
import Log
        

class ReadData:
    """    
    Class:
        商品情報の読み込み
    """    
    def readPriceCutProduct(self, lstPriceCutProduct):
        """
        Details:
            値下げ商品取得
        Args:
            lstPriceCutProduct  値下げ商品情報
        Return:
            lstPriceCutProduct  値下げ商品情報
        """
        logging = Log.getLogger()
        objCellData = Const.CellData()
        
        try:
            baseFilePath = os.path.dirname(os.path.abspath(sys.argv[0]))    # ×：__file__  / ○：sys.argv[0]
            strFilePath = os.path.normpath(os.path.join(baseFilePath, r'..\doc\ListingList.xlsx'))
            xlBook = openpyxl.load_workbook(strFilePath)
            xlSheet = xlBook.worksheets[0]
            
            nProductCount = 1
            try:
                
                for row in xlSheet.iter_rows(min_row=2):
                    # セルの値をリストとして取得
                    xlValues = [cell.value for cell in row]
                    
                    # 商品IDが空なら次へ
                    if xlValues[objCellData.PRODUCT_ID] is None:
                        break
                    
                    strProductName  = xlValues[objCellData.PRODUCT_NAME]       # 商品名
                    strProductId    = xlValues[objCellData.PRODUCT_ID]         # 商品ID
                    nCheapestPrice  = xlValues[objCellData.CHEAPES_PRICE]      # 最安値
                    
                    # 値下げ商品情報
                    objPriceCutInfo = ProductInfo.PriceCutProduct(strProductName, strProductId)
                    
                    # 最安値が「空白」でなければ
                    if nCheapestPrice is not None:
                        objPriceCutInfo.setCheapestPrice(nCheapestPrice)
                        
                    # 数値以外であれば警告を出し「0円」で設定
                    else:
                        objPriceCutInfo.setCheapestPrice(0)
                        logging.warning('最安値が「空白」または「文字列」で設定されています')
                        logging.warning(f'最安値を「0円」で設定し直しました。商品ID：{strProductId}')
                    
                    # 商品情報の追加
                    lstPriceCutProduct.append(objPriceCutInfo)
                    
                    logging.info(f'{nProductCount}つ目の商品情報を取得しました。商品ID：{strProductId}')
                    nProductCount += 1      # 商品番号のインクリメント
                    
            except Exception as e:
                nProductCount += 1      # 商品番号のインクリメント
                logging.error(f'{nProductCount}つ目の商品情報の取得に失敗しました')
                print(f'ExceptionLog : {e}')
                pass
            
        except FileNotFoundError as e:
            logging.critical('.xlsxの取得に失敗しました')
            print(f'ExceptionLog : {e}')
            
            
        return lstPriceCutProduct
    
    
    
class WriteData:
    """
    Class:
        商品情報の書き込み
    """
    def writeProductInfo(self, lstWriteProductInfo):
        """
        Details:
            値下げ商品取得
        Args:
            lstWriteProductInfo  値下げ商品情報
        """
        logging = Log.getLogger()
        objCellData = Const.CellData()
        
        # データファイルに記録されてある商品情報と出品中の商品情報に差異がない場合、追加記録せず終了
        if len(lstWriteProductInfo) == 0:
            logging.info('データファイルと出品中の商品情報に差異がないため追加記録は行いません')
            return
        
        try:
            baseFilePath = os.path.dirname(os.path.abspath(sys.argv[0]))    # ×：__file__  / ○：sys.argv[0]
            strFilePath = os.path.normpath(os.path.join(baseFilePath, r'..\doc\ListingList.xlsx'))
            xlBook = openpyxl.load_workbook(strFilePath)
            xlSheet = xlBook.worksheets[0]
            xlMaxRow = xlSheet.max_row      # 全範囲での最終行
            xlLastRow = 1
            
            nProductCount = 1
            # 最終行から逆ループ
            for row in reversed(range(1, xlMaxRow)):
                # 指定列の行が空でなくなるまで遡り最終的な指定列の最終行を探す
                if xlSheet.cell(row=row, column=objCellData.PRODUCT_ID + 1).value != None:
                    # 指定列の最終行を取得
                    xlLastRow = row + 1
                    break
                    
            try:
                # 最終行から新規商品情報を記載
                for objWriteProductInfo in lstWriteProductInfo:
                    xlSheet.cell(xlLastRow, objCellData.PRODUCT_NAME + 1, value=objWriteProductInfo.getProductName())     # 商品名
                    xlSheet.cell(xlLastRow, objCellData.PRODUCT_ID + 1, value=objWriteProductInfo.getProductId())         # 商品ID
                    
                    logging.info(f'{nProductCount}つ目の商品情報を記録しました。商品ID：{objWriteProductInfo.getProductId()}')
                    xlLastRow += 1          # 書き込み行のインクリメント
                    nProductCount += 1      # 商品番号のインクリメント
                    
            except Exception as e:
                nProductCount += 1      # 商品番号のインクリメント
                logging.error(f'{nProductCount}つ目の商品情報の記録に失敗しました')
                print(f'ExceptionLog : {e}')
                pass
                
                
            # 上書き保存
            xlBook.save(strFilePath)
            
        except Exception as e:
            logging.critical('.xlsxの取得に失敗しました')
            print(f'ExceptionLog : {e}')
        
        
        
    def writeProductStatus(self, lstWriteProductStatus):
        """
        Details:
            値下げ商品取得
        Args:
            lstWriteProductStatus   商品状態リスト（公開停止、削除、売り切れ）
        """
        logging = Log.getLogger()
        objCellData = Const.CellData()
        
        # データファイルの情報と値下げを行った際に得た商品状態に変更がない場合、追加記録せず終了
        if len(lstWriteProductStatus) == 0:
            logging.info('商品状態の新たな変更が確認できないため追加記録は行いません')
            return
    
        try:
            baseFilePath = os.path.dirname(os.path.abspath(sys.argv[0]))    # ×：__file__  / ○：sys.argv[0]
            strFilePath = os.path.normpath(os.path.join(baseFilePath, r'..\doc\ListingList.xlsx'))
            xlBook = openpyxl.load_workbook(strFilePath)
            xlSheet = xlBook.worksheets[0]

            try:
                # 範囲データを順次処理
                for tupleTargetCellData in xlSheet.iter_cols(min_row=2, min_col=objCellData.PRODUCT_ID + 1, max_col=objCellData.PRODUCT_ID + 1):
                    # タプルで返ってきた商品IDのデータを分析
                    for xlCellData in tupleTargetCellData:
                        
                        # 商品IDの取得
                        strCellData = xlCellData.value
                        
                        # 商品状態「公開停止、削除、売り切れ」のいずれかに該当する商品IDが存在するか範囲データ内を全検索
                        for objWriteProductStatus in lstWriteProductStatus:
                            
                            #「公開停止、削除、売り切れ」いずれかに該当がある商品IDがデータ内に存在するか否か
                            if strCellData == objWriteProductStatus.getProductId():
                                
                                # 「備考」セル位置を取得
                                xlCellData.offset(0, 2)
                                
                                if objWriteProductStatus.getProductStopPublishing():
                                    xlCellData.offset(0, 2).value = '公開停止'
                                    logging.info(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「公開停止」を記録しました')
                                    
                                if objWriteProductStatus.getProductDelete():
                                    xlCellData.offset(0, 2).value = '削除'
                                    logging.info(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「削除」を記録しました')
                                    
                                if objWriteProductStatus.getProductSoldOut():
                                    xlCellData.offset(0, 2).value = '売り切れ'
                                    logging.info(f'商品ID：{objWriteProductStatus.getProductId()} の備考欄に「売り切れ」を記録しました')
                                    
            except Exception as e:
                logging.error(f'商品状態の更新に失敗しました')
                print(f'ExceptionLog : {e}')
                pass
                
            # 上書き保存
            xlBook.save(strFilePath)
            logging.info('全商品状態の更新が完了しました')
                    
        except FileNotFoundError as e:
            logging.critical('.xlsxの取得に失敗しました')
            print(f'ExceptionLog : {e}')